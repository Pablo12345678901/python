import openpyxl

def Ajout_donnees_Excel_ds_dictionnaire(dictionnaire, fichier_excel):
    wb = openpyxl.load_workbook(fichier_excel, data_only=True)
    # chargement des data du fichier excel dans la variable wb
    sheet_wb = wb[wb.sheetnames[0]]
    # sélection de la première sheet du fichier excel
    for row in range(2, sheet_wb.max_row):
        # boucle sur les lignes 2 à ligne max
        nom_article = sheet_wb.cell(row,1).value
        # sélection de l'article en colonne 1
        if not nom_article:
            break
            # si la valeur la colonne 1 est nulle on break la boucle
        total_vente_par_article = sheet_wb.cell(row,4).value
        # prise de la valeur du total des ventes par article en colonne 4
        if nom_article not in dictionnaire:
            # si l'article n'est pas inclut dans le dictionnaire alors on l'ajoute
            dictionnaire[nom_article] = [total_vente_par_article]
            # Exemple : # {"Pommes" : [760]}
        # Sinon, on ajoute la valeur du total des ventes à l'article déjà existant
        # Version avec le total des ventes sous forme de int
        #else:
            #dictionnaire[nom_article][0] += total_vente_par_article
            # Exemple : # {"Pommes" : [760+840 = 1600]}
            # Donnera : # {"Pommes" : [1600]}
        # Autre version avec le total des ventes sous formes de liste :
        else:
            dictionnaire[nom_article].append(total_vente_par_article)
        # la clef nom_article correspond au total des ventes
        # création d'une liste qui contiendra les valeurs du total des ventes
    return dictionnaire

donnees = {}
# création d'un dictionnaire vide
liste_fichier_excel = ["octobre.xlsx","novembre.xlsx","decembre.xlsx"]
for fichier_excel in liste_fichier_excel:
    donnees = Ajout_donnees_Excel_ds_dictionnaire(donnees,fichier_excel)
    # Ajout des données de chaque fichier excel au dictionnaire
print(donnees)

wb_sortie = openpyxl.Workbook()
# création d'un nouveau workbook en mémoire
# ensuite il sera transformé en fichier Excel
sheet = wb_sortie.active
# création d'une fichier par défaut
liste_entete_colonne_du_tableau = ["Article"]
for fichier_excel in liste_fichier_excel:
    liste_entete_colonne_du_tableau.append(fichier_excel.rstrip(".xlsx"))
# Ajout du nom de chaque fichier excel à la liste des entêtes
#DEBUG print(liste_entete_colonne_du_tableau)
for i in range (len(liste_entete_colonne_du_tableau)):
    cell_name = "A"
    sheet[cell_name+str(i+1)] = liste_entete_colonne_du_tableau[i]
wb_sortie.save("total_vente_trimestre.xlsx")

