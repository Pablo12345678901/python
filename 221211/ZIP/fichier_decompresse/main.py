import openpyxl

def Ajout_donnees_Excel_ds_dictionnaire(dictionnaire, fichier_excel):
    wb = openpyxl.load_workbook(fichier_excel, data_only=True)
    # chargement des data du fichier excel dans la variable wb
    sheet_wb = wb[wb.sheetnames[0]]
    # sélection de la première sheet du fichier excel
    for row in range(2, sheet_wb.max_row):
        # boucle sur les lignes 2 à ligne max
        nom_article = sheet_wb.cell(row,1).value
        # sélection de l'article en colonne 1
        if not nom_article:
            break
            # si la valeur la colonne 1 est nulle on break la boucle
        total_vente_par_article = sheet_wb.cell(row,4).value
        # prise de la valeur du total des ventes par article en colonne 4
        if nom_article not in dictionnaire:
            # si l'article n'est pas inclut dans le dictionnaire alors on l'ajoute
            dictionnaire[nom_article] = [total_vente_par_article]
            # Exemple : # {"Pommes" : [760]}
        # Sinon, on ajoute la valeur du total des ventes à l'article déjà existant
        # Version avec le total des ventes sous forme de int
        #else:
            #dictionnaire[nom_article][0] += total_vente_par_article
            # Exemple : # {"Pommes" : [760+840 = 1600]}
            # Donnera : # {"Pommes" : [1600]}
        # Autre version avec le total des ventes sous formes de liste :
        else:
            dictionnaire[nom_article].append(total_vente_par_article)
        # la clef nom_article correspond au total des ventes
        # création d'une liste qui contiendra les valeurs du total des ventes
    return dictionnaire

donnees = {}
# création d'un dictionnaire vide
liste_fichier_excel = ["octobre.xlsx","novembre.xlsx","decembre.xlsx"]
for fichier_excel in liste_fichier_excel:
    donnees = Ajout_donnees_Excel_ds_dictionnaire(donnees,fichier_excel)
    # Ajout des données de chaque fichier excel au dictionnaire
#DEBUG print(donnees)

wb_sortie = openpyxl.Workbook()
# création d'un nouveau workbook en mémoire
# ensuite il sera transformé en fichier Excel
sheet = wb_sortie.active
# création d'une fichier par défaut
liste_entete_colonne_du_tableau = ["Article"]
# ajout de la valeur Article en première position dans la liste des entêtes
for fichier_excel in liste_fichier_excel:
    liste_entete_colonne_du_tableau.append(fichier_excel.rstrip(".xlsx"))
    # ajout de chaque nom de fichier sans les extensions dans la liste des entêtes
liste_entete_colonne_du_tableau.append("Total")
# ajout de la dernière valeur "Total" en dernière position dans la liste des entêtes
colonne = 1
# début à la colonne 1
for i in range (len(liste_entete_colonne_du_tableau)):
    # pour chaque entête de la liste
    sheet.cell(1,colonne).value = liste_entete_colonne_du_tableau[i]
    # attribution en tant que valeur de la cellule
    colonne += 1
    # décallage d'une colonne à chaque fois

ligne = 2
# début à la deuxième ligne
for i in donnees.items():
    colonne = 1
    # remettre la valeur de colonne à 1
    total = 0
    # création d'une variable total et mise à 0 à chaque boucle
    sheet.cell(ligne,colonne).value = i[0]
    # assignation de la clef du dictionnaire à la cellule colonne 1
    for y in range(len(i[1])):
        # pour chaque élément des valeurs assignées aux clef du dictionnaire
        colonne += 1
        # décallage d'une colonne
        sheet.cell(ligne, colonne).value = i[1][y]
        # assignation de l'élément à la cellule
        total += i[1][y]
        # incrémentation du total
    colonne += 1
    # décallage d'une colonne
    sheet.cell(ligne, colonne).value = total
    # assignation du total à la cellule tout à droite de la linge
    ligne += 1
    # passage à la ligne suivante
    # boucle ...


#CREATION DE GRAPHIQUE
# Toujours mettre avant le save SINON fonctionnera pas
# Besoin de créer 3 objets :
#   la référence (la plage des données)
#   la Série (ce à quoi correspondent les données = la légende du graphique)
#   le type de graphique

#ASSIGNATION DE LA REFERENCE
chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col = sheet.max_column, max_row=2)
# besoin de donner en paramètres la sheet, la colonne min, la ligne min, la colonne max, la ligne max

#ASSIGNATION DE LA SERIE = LA LEGENDE DES DONNEES DU GRAPHIQUE
chart_serie = openpyxl.chart.Series(chart_ref, title="Total Ventes CHF")
# besoin de donner en paramètres : la référence et le titre du graphique

#DEFINITION DU TYPE DE GRAPHIQUE
# Plus d'info sur les types ici : https://openpyxl.readthedocs.io/en/stable/charts/introduction.html
chart = openpyxl.chart.BarChart()
# création du chart de type Bar

# DÉFINITION DU TITRE DU GRAPHIQUE (CELUI D'EN HAUT)
chart.title = "Évolution du prix des pommes"

# AJOUT DE LA LEGENDE DES DONNEES AU GRAPHIQUE
chart.append(chart_serie)

# AJOUT DU GRAPHIQUE AU FICHIER EXCEL
sheet.add_chart(chart, "G2")
# besoin de fournir en paramètres le chart et la cellule ou il commencera

wb_sortie.save("total_ventes_bis.xlsx")
# sauvegarde du fichier sous un nouveau nom

str = 'utiliser un nombre de paramètres illimités dans une fonction avec "*":'
print(str.upper())


